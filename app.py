# app.py — Flask メインアプリケーション
# キリン屋 予約管理システム
from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, Response, session
import json, socket, time as _time, csv, io, os, base64
from datetime import datetime, date, timedelta
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from database import (
    init_db, save_reservation, get_reservations_by_date,
    get_reservation, update_reservation, cancel_reservation,
    set_reservation_status, get_all_reservations,
    get_cancelled_reservations_by_date, engine,
    apply_pos_sale_to_reservation, insert_pos_sales,
    get_pos_sales, set_pos_sale_group_match,
    get_deleted_reservations, soft_delete_reservation, restore_reservation,
)
from logic import (
    assign_seats, check_time_conflict, get_table_status, get_seat_map_status,
    get_customer_history, aggregate_customer_ranking, customer_key,
    match_pos_sale_groups, TABLES
)
import pos_import as posimp

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'kiriniya-dev-only-change-in-prod')

APP_PASSWORD = os.environ.get('APP_PASSWORD', '')

# ============================================================
# DB初期化（モジュール読み込み時に実行）
# gunicornは `if __name__ == '__main__':` を通らず app を直接読み込むため、
# ここで呼ばないと本番環境（Render等）でテーブルが作成されない
# ============================================================
try:
    init_db()
except Exception as e:
    print(f"[WARN] DB初期化に失敗しました: {e}")

@app.before_request
def require_auth():
    public = {'login', 'static', 'api_last_change'}
    if not APP_PASSWORD or request.endpoint in public or session.get('ok'):
        return
    return redirect(url_for('login', next=request.path))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        if request.form.get('pw') == APP_PASSWORD:
            session['ok'] = True
            return redirect(request.args.get('next') or url_for('index'))
        return render_template('login.html', error=True)
    return render_template('login.html', error=False)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ============================================================
# リアルタイム同期サポート
# ============================================================
_CHANGE_FILE = 'last_change.txt'

def _notify_change(date_str: str):
    """予約変更をファイルに記録 → 他端末がポーリングで検知する"""
    try:
        with open(_CHANGE_FILE, 'w', encoding='utf-8') as f:
            f.write(f"{date_str}|{_time.time()}")
    except Exception:
        pass

def _get_local_ip() -> str:
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return '127.0.0.1'

LOCAL_IP = _get_local_ip()
app.jinja_env.globals['local_ip'] = LOCAL_IP

# ============================================================
# ヘルパー
# ============================================================
WEEKDAYS_JP = ['月', '火', '水', '木', '金', '土', '日']

def jp_date(dt_obj) -> str:
    """date → '2026年5月26日（火）'"""
    if isinstance(dt_obj, str):
        dt_obj = datetime.strptime(dt_obj, '%Y-%m-%d').date()
    return f"{dt_obj.year}年{dt_obj.month}月{dt_obj.day}日（{WEEKDAYS_JP[dt_obj.weekday()]}）"

def get_time_slots():
    """11:00〜22:00 を15分刻みで返す"""
    slots = []
    cur = datetime.strptime('11:00', '%H:%M')
    end = datetime.strptime('22:00', '%H:%M')
    while cur <= end:
        slots.append(cur.strftime('%H:%M'))
        cur += timedelta(minutes=15)
    return slots

DURATION_OPTIONS = [
    (90,  '1時間30分'),
    (105, '1時間45分（標準）'),
    (120, '2時間'),
    (150, '2時間30分'),
    (180, '3時間'),
]

# ============================================================
# 特記事項タグ（現場スタッフが一目で配慮点を把握できるようにする）
# 'kids'（お子様連れ）はチェックボックスで手入力せず、children_info の
# 有無から自動判定する（二重入力・矛盾を防ぐため）
# ============================================================
SPECIAL_TAGS = {
    'kids':         {'icon': '👶', 'label': 'お子様連れ',       'color': '#6f42c1'},
    'elderly':      {'icon': '🧓', 'label': 'ご高齢のお客様',    'color': '#6c757d'},
    'pregnant':     {'icon': '🤰', 'label': '妊婦さんあり',      'color': '#d63384'},
    'birthday':     {'icon': '🎂', 'label': 'バースデー・記念日', 'color': '#fd7e14'},
    'allergy':      {'icon': '⚠️', 'label': 'アレルギーあり',    'color': '#dc3545'},
    'seat_request': {'icon': '🪑', 'label': 'お席のご希望あり',   'color': '#0d6efd'},
}
SPECIAL_TAG_CHOICES = ['elderly', 'pregnant', 'birthday', 'allergy', 'seat_request']  # フォームで選択可能なタグ

def active_special_tags(res: dict) -> list:
    """予約に紐づく特記タグのキー一覧（'kids' は children_info から自動付与）"""
    tags = []
    if res.get('children_info'):
        tags.append('kids')
    for t in (res.get('special_tags') or []):
        if t in SPECIAL_TAG_CHOICES:
            tags.append(t)
    return tags

# ============================================================
# 属性別スタッフ準備指示
# 子供の年齢・妊婦・高齢者タグから、現場で必要な準備をルールベースで導出する
# ============================================================
def staff_prep_instructions(res: dict) -> list:
    out = []
    ages = [c.get('age') for c in (res.get('children_info') or []) if isinstance(c, dict) and c.get('age') is not None]

    if any(a == 0 for a in ages):
        out.append({'icon': '🛏️', 'text': 'ベビーベッドの用意', 'reason': '0歳の赤ちゃんがいます'})
    if any(a <= 12 for a in ages):
        out.append({'icon': '🍶', 'text': '子供のタレを用意', 'reason': '小学生以下のお子様がいます'})
    if any(a <= 3 for a in ages):
        out.append({'icon': '🍴', 'text': 'カトラリー・小さい取り皿・スプーン・フォークを用意', 'reason': '3歳以下のお子様がいます'})

    tags = res.get('special_tags') or []
    if 'pregnant' in tags:
        out.append({'icon': '🪑', 'text': '座布団（クッションなど）が必要', 'reason': '妊婦様がいらっしゃいます'})
        out.append({'icon': '🍚', 'text': '安産米が必要', 'reason': '妊婦様がいらっしゃいます'})
    if 'elderly' in tags:
        out.append({'icon': '💺', 'text': '座椅子が必要かを確認する', 'reason': 'ご高齢のお客様がいらっしゃいます'})

    return out

# ============================================================
# 常連マーク（手動フラグ is_regular OR 来店回数による自動判定）
# ============================================================
REGULAR_VISIT_THRESHOLD = 3  # 累計来店（予約）回数がこれ以上で自動的に「常連」扱い

def _mark_regular_customers(reservations: list) -> list:
    """
    予約リストの各要素に is_regular_effective（手動フラグ or 自動判定）を付与する。
    リストは in-place で変更され、そのまま戻り値としても返す。
    """
    all_res = get_all_reservations(include_cancelled=False)
    ranking = aggregate_customer_ranking(all_res)
    regular_keys = {c['key'] for c in ranking if c['visit_count'] >= REGULAR_VISIT_THRESHOLD}
    for r in reservations:
        r['is_regular_effective'] = bool(r.get('is_regular')) or (customer_key(r) in regular_keys)
    return reservations

# Jinja2グローバル関数
app.jinja_env.globals['jp_date'] = jp_date
app.jinja_env.globals['SPECIAL_TAGS'] = SPECIAL_TAGS
app.jinja_env.globals['SPECIAL_TAG_CHOICES'] = SPECIAL_TAG_CHOICES
app.jinja_env.globals['active_special_tags'] = active_special_tags
app.jinja_env.globals['staff_prep_instructions'] = staff_prep_instructions
app.jinja_env.globals['REGULAR_VISIT_THRESHOLD'] = REGULAR_VISIT_THRESHOLD

# ============================================================
# ルート: トップ（日次一覧）
# ============================================================
@app.route('/')
def index():
    target_date = request.args.get('date', date.today().isoformat())
    try:
        target_dt = datetime.strptime(target_date, '%Y-%m-%d').date()
    except ValueError:
        target_dt   = date.today()
        target_date = target_dt.isoformat()

    reservations = get_reservations_by_date(target_date)
    _mark_regular_customers(reservations)
    table_status = get_table_status(reservations)
    cancelled    = get_cancelled_reservations_by_date(target_date)

    return render_template(
        'index.html',
        reservations = reservations,
        cancelled    = cancelled,
        target_date  = target_date,
        target_dt    = target_dt,
        date_display = jp_date(target_dt),
        prev_date    = (target_dt - timedelta(days=1)).isoformat(),
        next_date    = (target_dt + timedelta(days=1)).isoformat(),
        table_status = table_status,
        tables       = TABLES,
        today        = date.today().isoformat(),
    )

# ============================================================
# ルート: 新規予約
# ============================================================
@app.route('/reservation/new', methods=['GET', 'POST'])
def new_reservation():
    if request.method == 'POST':
        res_dict, err = _parse_form(request)
        if err:
            flash(f'⚠️ 入力エラー: {err}', 'danger')
            return redirect(request.url)

        existing = get_reservations_by_date(res_dict['date'])

        # 入店時間競合チェック
        conflicts = check_time_conflict(res_dict, existing)
        if conflicts and not res_dict['is_vip']:
            names = '、'.join(r['name'] for r in conflicts)
            flash(f"⚠️ 入店時間警告: {names} 様と15分以内の入店になります（VIPフラグでスキップ可）", 'warning')

        # 席の確定（スタッフが席マップで選択した場合はそれを優先）
        confirmed = _parse_confirmed_tables(request.form.get('confirmed_tables', '[]'))
        if confirmed:
            res_dict['assigned_tables'] = json.dumps(confirmed)
            seat_note = '・'.join(str(t) + '番' for t in confirmed) + '（スタッフ確定）'
        else:
            # 未選択の場合はシステム提案を自動使用
            suggestion = assign_seats(res_dict, existing)
            res_dict['assigned_tables'] = json.dumps(suggestion['tables'])
            seat_note = suggestion['note'] + '（自動）'

        save_reservation(res_dict)
        flash(f"✅ 予約登録完了 ｜ {res_dict['name']} 様 ／ {seat_note}", 'success')
        return redirect(url_for('index', date=res_dict['date']))

    default_date = request.args.get('date', date.today().isoformat())
    return render_template(
        'form.html',
        mode             = 'new',
        reservation      = None,
        time_slots       = get_time_slots(),
        duration_options = DURATION_OPTIONS,
        default_date     = default_date,
        tables           = TABLES,
    )

# ============================================================
# ルート: 予約詳細
# ============================================================
@app.route('/reservation/<int:rid>')
def view_reservation(rid):
    res = get_reservation(rid)
    if not res:
        flash('予約が見つかりません', 'danger')
        return redirect(url_for('index'))
    if res.get('is_deleted'):
        flash('この予約はゴミ箱に移動されています。復元してから確認してください', 'warning')
        return redirect(url_for('trash_page'))
    all_res = get_all_reservations(include_cancelled=False)
    history = get_customer_history(res['name'], res['phone'], all_res, exclude_id=rid)
    # 自分自身を含めた累計来店回数で常連判定（他画面の判定基準と統一）
    res['is_regular_effective'] = bool(res.get('is_regular')) or (history['visit_count'] + 1 >= REGULAR_VISIT_THRESHOLD)
    return render_template('detail.html', reservation=res, tables=TABLES, history=history)

# ============================================================
# ルート: 予約編集
# ============================================================
@app.route('/reservation/<int:rid>/edit', methods=['GET', 'POST'])
def edit_reservation(rid):
    res = get_reservation(rid)
    if not res:
        flash('予約が見つかりません', 'danger')
        return redirect(url_for('index'))
    if res.get('is_deleted'):
        flash('この予約はゴミ箱に移動されています。復元してから編集してください', 'warning')
        return redirect(url_for('trash_page'))

    if request.method == 'POST':
        updated, err = _parse_form(request, edit_id=rid)
        if err:
            flash(f'⚠️ 入力エラー: {err}', 'danger')
            return redirect(request.url)

        # 自分以外の予約で競合チェック
        existing = [r for r in get_reservations_by_date(updated['date']) if r['id'] != rid]
        conflicts = check_time_conflict(updated, existing)
        if conflicts and not updated['is_vip']:
            names = '、'.join(r['name'] for r in conflicts)
            flash(f"⚠️ 入店時間警告: {names} 様と15分以内の入店になります", 'warning')

        # 席の確定（スタッフ選択 or 自動提案）
        confirmed = _parse_confirmed_tables(request.form.get('confirmed_tables', '[]'))
        if confirmed:
            updated['assigned_tables'] = json.dumps(confirmed)
            seat_note = '・'.join(str(t) + '番' for t in confirmed) + '（スタッフ確定）'
        else:
            suggestion = assign_seats(updated, existing)
            updated['assigned_tables'] = json.dumps(suggestion['tables'])
            seat_note = suggestion['note'] + '（自動）'
        updated['status'] = res['status']

        update_reservation(rid, updated)
        flash(f"✅ 予約を更新しました ｜ {seat_note}", 'success')
        return redirect(url_for('view_reservation', rid=rid))

    return render_template(
        'form.html',
        mode             = 'edit',
        reservation      = res,
        time_slots       = get_time_slots(),
        duration_options = DURATION_OPTIONS,
        default_date     = res['date'],
        tables           = TABLES,
    )

# ============================================================
# ルート: キャンセル
# ============================================================
@app.route('/reservation/<int:rid>/cancel', methods=['POST'])
def cancel_route(rid):
    res = get_reservation(rid)
    if res:
        cancel_reservation(rid)
        flash(f"🗑️ {res['name']} 様の予約をキャンセルしました", 'info')
        return redirect(url_for('index', date=res['date']))
    return redirect(url_for('index'))

# ============================================================
# ルート: 来店受付（着席）トグル
# ============================================================
@app.route('/reservation/<int:rid>/toggle-visited', methods=['POST'])
def toggle_visited_route(rid):
    """来店済み ⇔ 未来店（confirmed）を切り替える。現場で着席が一目でわかるようにする。"""
    res = get_reservation(rid)
    if not res:
        flash('予約が見つかりません', 'danger')
        return redirect(url_for('index'))
    if res['status'] == 'cancelled':
        flash('キャンセル済みの予約は来店受付できません', 'warning')
        return redirect(request.referrer or url_for('index', date=res['date']))

    new_status = 'confirmed' if res['status'] == 'visited' else 'visited'
    set_reservation_status(rid, new_status)
    if new_status == 'visited':
        flash(f"🪑 {res['name']} 様 来店受付しました（着席）", 'success')
    else:
        flash(f"↩️ {res['name']} 様の来店ステータスを解除しました", 'info')
    return redirect(request.referrer or url_for('index', date=res['date']))

# ============================================================
# ルート: 論理削除（ゴミ箱）
# ============================================================
@app.route('/reservation/<int:rid>/delete', methods=['POST'])
def delete_reservation_route(rid):
    """
    予約をゴミ箱へ移動する。DBから完全に消去することはなく、
    is_deleted フラグを立てるのみなので /trash からいつでも復元できる。
    """
    res = get_reservation(rid)
    if not res:
        flash('予約が見つかりません', 'danger')
        return redirect(url_for('index'))
    soft_delete_reservation(rid)
    flash(f"🗑️ {res['name']} 様の予約をゴミ箱へ移動しました（復元は「ゴミ箱」から可能です）", 'info')
    return redirect(url_for('index', date=res['date']))


@app.route('/trash')
def trash_page():
    deleted = get_deleted_reservations()
    return render_template('trash.html', deleted=deleted)


@app.route('/reservation/<int:rid>/restore', methods=['POST'])
def restore_reservation_route(rid):
    res = get_reservation(rid)
    if not res:
        flash('予約が見つかりません', 'danger')
        return redirect(url_for('trash_page'))
    restore_reservation(rid)
    flash(f"✅ {res['name']} 様の予約を復元しました", 'success')
    return redirect(url_for('trash_page'))

# ============================================================
# ルート: タイムテーブル
# ============================================================
OPEN_HOUR   = 11    # タイムテーブル表示開始
CLOSE_HOUR  = 23    # タイムテーブル表示終了
PX_PER_MIN  = 2     # 通常タイム 1分あたりのピクセル数
ROTATION_WARN_GAP_DANGER  = 15
ROTATION_WARN_GAP_CAUTION = 30

# インターバル圧縮（ランチ〜ディナーの準備時間を短縮表示）
LUNCH_END_HOUR      = 14    # ランチ閉店・インターバル開始
DINNER_START_HOUR   = 17    # ディナー開店・インターバル終了
BREAK_COMPRESSED_PX = 40    # インターバルの表示高さ（px）

# スケジュールマーカー定義（開店・OS・閉店ライン）
SCHEDULE_MARKERS = [
    {'label': 'ランチ開店', 'hour': 11, 'min': 30, 'color': '#2ec4b6', 'dashed': False, 'below': False},
    {'label': 'ランチOS',   'hour': 13, 'min': 30, 'color': '#ff9800', 'dashed': True,  'below': False},
    {'label': 'ディナーOS', 'hour': 22, 'min': 0,  'color': '#ff9800', 'dashed': True,  'below': False},
    {'label': 'ディナー閉店','hour': 23, 'min': 0,  'color': '#e63946', 'dashed': False, 'below': False},
]

@app.route('/timetable')
def timetable():
    target_date = request.args.get('date', date.today().isoformat())
    try:
        target_dt = datetime.strptime(target_date, '%Y-%m-%d').date()
    except ValueError:
        target_dt   = date.today()
        target_date = target_dt.isoformat()

    reservations = get_reservations_by_date(target_date)
    _mark_regular_customers(reservations)

    # ── インターバル圧縮 px変換 ──
    lunch_end_min    = (LUNCH_END_HOUR   - OPEN_HOUR) * 60   # 180
    dinner_start_min = (DINNER_START_HOUR - OPEN_HOUR) * 60  # 360
    break_duration   = dinner_start_min - lunch_end_min       # 180
    break_start_px   = lunch_end_min * PX_PER_MIN             # 360
    break_end_px     = break_start_px + BREAK_COMPRESSED_PX   # 400

    def min_to_px(m):
        if m <= lunch_end_min:
            return round(m * PX_PER_MIN)
        elif m <= dinner_start_min:
            t = (m - lunch_end_min) / break_duration
            return round(break_start_px + t * BREAK_COMPRESSED_PX)
        else:
            return round(break_end_px + (m - dinner_start_min) * PX_PER_MIN)

    total_minutes = (CLOSE_HOUR - OPEN_HOUR) * 60
    total_height  = min_to_px(total_minutes)

    # ── テーブルごとに回転情報を集計 ──
    table_data = {}
    for table_id in TABLES:
        res_list = [r for r in reservations if table_id in r.get('assigned_tables', [])]
        res_list.sort(key=lambda r: r['time_slot'])

        rotations = []
        for i, res in enumerate(res_list):
            sh, sm_h = map(int, res['time_slot'].split(':'))
            start_min = (sh - OPEN_HOUR) * 60 + sm_h
            end_min   = start_min + int(res['duration_minutes'])
            top_px    = min_to_px(start_min)
            height_px = max(min_to_px(end_min) - top_px, 36)

            # 直前回転との間隔
            gap_min = None
            if i > 0:
                prev = rotations[-1]
                gap_min = start_min - prev['_end_min']

            rotation_num = i + 1
            gap_danger  = gap_min is not None and 0 <= gap_min < ROTATION_WARN_GAP_DANGER
            gap_caution = gap_min is not None and ROTATION_WARN_GAP_DANGER <= gap_min < ROTATION_WARN_GAP_CAUTION
            rotations.append({
                **res,
                'rotation_num': rotation_num,
                'top_px':       top_px,
                'height_px':    height_px,
                'gap_min':      gap_min,
                'warn_3rd':     rotation_num >= 3,
                'warn_gap':     gap_danger or gap_caution,
                'gap_danger':   gap_danger,
                'gap_caution':  gap_caution,
                '_end_min':     end_min,
            })
        table_data[table_id] = rotations

    # ── 複数テーブル連結ブリッジ（団体様が隣接テーブルにまたがる場合、
    #    タイムテーブル上で視覚的につなげて描画するための座標データ）──
    col_index = {t_id: i for i, t_id in enumerate(TABLES.keys())}
    group_bridges = []
    for res in reservations:
        tbls = res.get('assigned_tables') or []
        if len(tbls) < 2:
            continue
        sorted_tbls = sorted((t for t in tbls if t in col_index), key=lambda t: col_index[t])
        sh, sm_h  = map(int, res['time_slot'].split(':'))
        start_min = (sh - OPEN_HOUR) * 60 + sm_h
        end_min   = start_min + int(res['duration_minutes'])
        top_px    = min_to_px(start_min)
        height_px = max(min_to_px(end_min) - top_px, 36)
        for a, b in zip(sorted_tbls, sorted_tbls[1:]):
            if col_index[b] - col_index[a] == 1:   # 隣接列同士のみブリッジ描画
                group_bridges.append({
                    'col_index': col_index[a],
                    'top_px':    top_px,
                    'height_px': height_px,
                })

    # ── 時刻ラベル（30分刻み、インターバル内省略）──
    time_labels = []
    for h in range(OPEN_HOUR, CLOSE_HOUR + 1):
        for mt in (0, 30):
            if h == CLOSE_HOUR and mt > 0:
                break
            from_open = (h - OPEN_HOUR) * 60 + mt
            if lunch_end_min < from_open < dinner_start_min:
                continue   # インターバル内はスキップ
            time_labels.append({
                'label':   f'{h:02d}:{mt:02d}',
                'top_px':  min_to_px(from_open),
                'is_hour': mt == 0,
            })

    # ── グリッド線（15分ごと、インターバル内省略）──
    grid_lines = []
    for slot in range(0, total_minutes, 15):
        if lunch_end_min < slot < dinner_start_min:
            continue
        grid_lines.append({
            'top_px':  min_to_px(slot),
            'is_hour': slot % 60 == 0,
        })

    # ── スケジュールマーカー（px位置を付加）──
    schedule_markers = []
    for sm_def in SCHEDULE_MARKERS:
        from_open = (sm_def['hour'] - OPEN_HOUR) * 60 + sm_def['min']
        schedule_markers.append({**sm_def, 'top_px': min_to_px(from_open)})

    # ── 警告リスト ──
    warnings = []
    for tid, rots in table_data.items():
        for r in rots:
            if r['warn_3rd']:
                warnings.append({
                    'type': '3rd',
                    'msg':  f"{tid}番テーブル ｜ {r['name']} 様 {r['time_slot']}〜 が3回転目です",
                })
            if r['gap_danger']:
                warnings.append({
                    'type': 'danger',
                    'msg':  f"{tid}番テーブル ｜ {r['name']} 様 {r['time_slot']}〜 の前の回転との間隔が {r['gap_min']} 分です（15分未満・要確認）",
                })
            elif r['gap_caution']:
                warnings.append({
                    'type': 'caution',
                    'msg':  f"{tid}番テーブル ｜ {r['name']} 様 {r['time_slot']}〜 の前の回転との間隔が {r['gap_min']} 分です（15〜30分・注意）",
                })

    return render_template(
        'timetable.html',
        target_date      = target_date,
        target_dt        = target_dt,
        date_display     = jp_date(target_dt),
        prev_date        = (target_dt - timedelta(days=1)).isoformat(),
        next_date        = (target_dt + timedelta(days=1)).isoformat(),
        today            = date.today().isoformat(),
        tables           = TABLES,
        table_data       = table_data,
        time_labels      = time_labels,
        grid_lines       = grid_lines,
        total_height     = total_height,
        total_minutes    = total_minutes,
        warnings         = warnings,
        px_per_min       = PX_PER_MIN,
        open_hour        = OPEN_HOUR,
        close_hour       = CLOSE_HOUR,
        reservations     = reservations,
        schedule_markers = schedule_markers,
        group_bridges    = group_bridges,
        break_start_px   = break_start_px,
        break_end_px     = break_end_px,
        lunch_end_min    = lunch_end_min,
        dinner_start_min = dinner_start_min,
    )

# ============================================================
# ルート: 印刷専用（ランチ・ディナー・サマリー 各1ページ）
# ============================================================
@app.route('/timetable/print')
def timetable_print_view():
    target_date = request.args.get('date', date.today().isoformat())
    try:
        target_dt = datetime.strptime(target_date, '%Y-%m-%d').date()
    except ValueError:
        target_dt = date.today(); target_date = target_dt.isoformat()

    reservations = get_reservations_by_date(target_date)

    lunch_end_min    = (LUNCH_END_HOUR   - OPEN_HOUR) * 60
    dinner_start_min = (DINNER_START_HOUR - OPEN_HOUR) * 60
    break_duration   = dinner_start_min - lunch_end_min
    break_start_px   = lunch_end_min * PX_PER_MIN
    break_end_px     = break_start_px + BREAK_COMPRESSED_PX

    def mtp(m):
        if m <= lunch_end_min:
            return round(m * PX_PER_MIN)
        elif m <= dinner_start_min:
            return round(break_start_px + (m - lunch_end_min) / break_duration * BREAK_COMPRESSED_PX)
        else:
            return round(break_end_px + (m - dinner_start_min) * PX_PER_MIN)

    total_minutes = (CLOSE_HOUR - OPEN_HOUR) * 60
    total_h       = mtp(total_minutes)

    # テーブルデータ（メインと同じ計算）
    table_data = {}
    for table_id in TABLES:
        res_list = [r for r in reservations if table_id in r.get('assigned_tables', [])]
        res_list.sort(key=lambda r: r['time_slot'])
        rotations = []
        for i, res in enumerate(res_list):
            sh, sm_h  = map(int, res['time_slot'].split(':'))
            start_min = (sh - OPEN_HOUR) * 60 + sm_h
            end_min   = start_min + int(res['duration_minutes'])
            top_px    = mtp(start_min)
            height_px = max(mtp(end_min) - top_px, 36)
            gap_min   = (start_min - rotations[-1]['_end_min']) if i > 0 else None
            rn        = i + 1
            gd        = gap_min is not None and 0 <= gap_min < ROTATION_WARN_GAP_DANGER
            gc        = gap_min is not None and ROTATION_WARN_GAP_DANGER <= gap_min < ROTATION_WARN_GAP_CAUTION
            rotations.append({**res, 'rotation_num': rn, 'top_px': top_px,
                               'height_px': height_px, 'gap_min': gap_min,
                               'warn_3rd': rn >= 3, 'gap_danger': gd, 'gap_caution': gc,
                               '_end_min': end_min})
        table_data[table_id] = rotations

    # ── ランチ（11:00〜14:00） ──
    lunch_h = break_start_px
    lunch_labels = [
        {'label': f'{h:02d}:{mt:02d}', 'top_px': mtp((h-OPEN_HOUR)*60+mt), 'is_hour': mt==0}
        for h in range(OPEN_HOUR, LUNCH_END_HOUR+1)
        for mt in (0, 30)
        if not (h == LUNCH_END_HOUR and mt > 0)
    ]
    lunch_grid = [{'top_px': mtp(s), 'is_hour': s%60==0}
                  for s in range(0, lunch_end_min+1, 15)]
    lunch_markers = [
        {'label': 'ランチ開店', 'top_px': mtp(30),    'color': '#2ec4b6', 'dashed': False, 'above': False},
        {'label': 'ランチOS',   'top_px': mtp(150),   'color': '#ff9800', 'dashed': True,  'above': True},
        {'label': 'ランチ閉店', 'top_px': lunch_h,    'color': '#e63946', 'dashed': False, 'above': True},
    ]
    lunch_tdata = {t: [r for r in rots if r['top_px'] < lunch_h]
                   for t, rots in table_data.items()}

    # ── ディナー（17:00〜23:00） ──
    off       = break_end_px
    dinner_h  = total_h - off
    dinner_labels = [
        {'label': f'{h:02d}:{mt:02d}', 'top_px': mtp((h-OPEN_HOUR)*60+mt)-off, 'is_hour': mt==0}
        for h in range(DINNER_START_HOUR, CLOSE_HOUR+1)
        for mt in (0, 30)
        if not (h == CLOSE_HOUR and mt > 0)
    ]
    dinner_grid = [{'top_px': mtp(s)-off, 'is_hour': s%60==0}
                   for s in range(dinner_start_min, total_minutes+1, 15)]
    dinner_markers = [
        {'label': 'ディナー開店', 'top_px': 0,                              'color': '#2ec4b6', 'dashed': False, 'above': False},
        {'label': 'ディナーOS',   'top_px': mtp((22-OPEN_HOUR)*60)-off,    'color': '#ff9800', 'dashed': True,  'above': True},
        {'label': 'ディナー閉店', 'top_px': dinner_h,                       'color': '#e63946', 'dashed': False, 'above': True},
    ]
    dinner_tdata = {
        t: [dict(r, top_px=r['top_px']-off) for r in rots if r['top_px'] >= off]
        for t, rots in table_data.items()
    }

    # 警告
    warnings = []
    for tid, rots in table_data.items():
        for r in rots:
            if r['warn_3rd']:
                warnings.append({'type': '3rd',     'msg': f"{tid}番 {r['name']} 様 {r['time_slot']}〜 が3回転目"})
            if r['gap_danger']:
                warnings.append({'type': 'danger',  'msg': f"{tid}番 {r['name']} 様 間隔{r['gap_min']}分（15分未満）"})
            elif r['gap_caution']:
                warnings.append({'type': 'caution', 'msg': f"{tid}番 {r['name']} 様 間隔{r['gap_min']}分（15〜30分）"})

    col_w = 94  # A4横(余白6mm・軸40px): 1078px - 40px軸 = 1038px ÷ 11列 ≈ 94px

    # ── 飛び込み客記入欄の行数（最大40組基準） ──
    MAX_GROUPS       = 40
    total_reserved   = len(reservations)
    walk_in_slots    = max(0, MAX_GROUPS - total_reserved)
    lunch_res_count  = sum(1 for r in reservations
                           if int(r['time_slot'].split(':')[0]) < LUNCH_END_HOUR)
    dinner_res_count = sum(1 for r in reservations
                           if int(r['time_slot'].split(':')[0]) >= DINNER_START_HOUR)
    lunch_walkin  = max(5, min(12, round(walk_in_slots * 0.4)))
    dinner_walkin = max(8, min(15, round(walk_in_slots * 0.6)))
    if lunch_walkin + dinner_walkin > 25:
        dinner_walkin = 25 - lunch_walkin

    return render_template('timetable_print.html',
        date_display     = jp_date(target_dt),
        tables           = TABLES,
        col_w            = col_w,
        lunch_h          = lunch_h,
        lunch_labels     = lunch_labels,
        lunch_grid       = lunch_grid,
        lunch_markers    = lunch_markers,
        lunch_tdata      = lunch_tdata,
        dinner_h         = dinner_h,
        dinner_labels    = dinner_labels,
        dinner_grid      = dinner_grid,
        dinner_markers   = dinner_markers,
        dinner_tdata     = dinner_tdata,
        table_data       = table_data,
        reservations     = reservations,
        warnings         = warnings,
        lunch_res_count  = lunch_res_count,
        dinner_res_count = dinner_res_count,
        lunch_walkin     = lunch_walkin,
        dinner_walkin    = dinner_walkin,
        max_groups       = MAX_GROUPS,
    )

# ============================================================
# ルート: 予約一覧 印刷専用（紙の予約台帳フォーマット・A4縦）
# ============================================================
@app.route('/reservations/print')
def reservations_print_view():
    target_date = request.args.get('date', date.today().isoformat())
    try:
        target_dt = datetime.strptime(target_date, '%Y-%m-%d').date()
    except ValueError:
        target_dt = date.today(); target_date = target_dt.isoformat()

    reservations = get_reservations_by_date(target_date)
    cancelled    = get_cancelled_reservations_by_date(target_date)
    total_guests = sum(r['total_people'] for r in reservations)

    # 手書き飛び込み客欄：最大40組を基準に残り枠数ぶん空行を用意
    MAX_GROUPS  = 40
    walkin_rows = max(6, MAX_GROUPS - len(reservations))

    return render_template('reservation_list_print.html',
        date_display  = jp_date(target_dt),
        reservations  = reservations,
        cancelled     = cancelled,
        total_guests  = total_guests,
        walkin_rows   = walkin_rows,
    )

# ============================================================
# 厨房用モニターモード（本日の状況をリアルタイム大画面表示）
# ============================================================
def _kitchen_monitor_data() -> dict:
    """厨房モニター画面が表示する本日の集計データを計算する。常に「今日」基準。"""
    today = date.today()
    today_str = today.isoformat()
    now = datetime.now()

    reservations = get_reservations_by_date(today_str)  # キャンセルは含まれない

    total_groups = len(reservations)
    total_people = sum(r['total_people'] for r in reservations)
    visited      = [r for r in reservations if r['status'] == 'visited']
    visited_groups = len(visited)
    visited_people = sum(r['total_people'] for r in visited)

    upcoming = [r for r in reservations if r['status'] != 'visited']
    upcoming.sort(key=lambda r: r['time_slot'])

    def _to_dt(r):
        h, m = map(int, r['time_slot'].split(':'))
        return datetime(today.year, today.month, today.day, h, m)

    next_res = None
    for r in upcoming:
        rt = _to_dt(r)
        if rt >= now:
            next_res = {
                'name':          r['name'],
                'time_slot':     r['time_slot'],
                'total_people':  r['total_people'],
                'minutes_until': round((rt - now).total_seconds() / 60),
            }
            break
    if next_res is None and upcoming:
        r = upcoming[0]
        rt = _to_dt(r)
        next_res = {
            'name':          r['name'],
            'time_slot':     r['time_slot'],
            'total_people':  r['total_people'],
            'minutes_until': round((rt - now).total_seconds() / 60),  # 負の値＝来店予定時刻超過
        }

    return {
        'date_display':    jp_date(today),
        'now':             now.strftime('%H:%M'),
        'total_groups':    total_groups,
        'total_people':    total_people,
        'visited_groups':  visited_groups,
        'visited_people':  visited_people,
        'remaining_groups': total_groups - visited_groups,
        'remaining_people': total_people - visited_people,
        'next':            next_res,
    }

@app.route('/kitchen-monitor')
def kitchen_monitor():
    return render_template('kitchen_monitor.html', data=_kitchen_monitor_data())

@app.route('/api/kitchen-monitor-data')
def api_kitchen_monitor_data():
    return jsonify(_kitchen_monitor_data())

# ============================================================
# API: リアルタイム競合チェック & 席提案
# ============================================================
@app.route('/api/check', methods=['POST'])
def api_check():
    """フォームからAJAXで呼ばれる。競合と席提案を返す。"""
    data = request.get_json() or {}
    target_date = data.get('date', '')
    time_slot   = data.get('time_slot', '')

    if not target_date or not time_slot:
        return jsonify({'conflicts': [], 'seat_suggestion': {'tables': [], 'note': '日付・時間を入力してください'}})

    existing = get_reservations_by_date(target_date)
    edit_id  = data.get('edit_id')
    if edit_id:
        existing = [r for r in existing if r['id'] != int(edit_id)]

    excl = int(edit_id) if edit_id else None
    conflicts  = check_time_conflict(data, existing)
    suggestion = assign_seats(data, existing)
    seat_map   = get_seat_map_status(
        time_slot,
        data.get('duration_minutes', 105),
        existing,
        excl
    )

    return jsonify({
        'conflicts':      [{'name': r['name'], 'time_slot': r['time_slot']} for r in conflicts],
        'seat_suggestion': suggestion,
        'seat_map':        seat_map,   # 使用中・ブロック中テーブルID一覧
    })

# ============================================================
# API: 予約1件取得（モーダル用）
# ============================================================
@app.route('/api/reservation/<int:rid>')
def api_get_reservation(rid):
    res = get_reservation(rid)
    if not res:
        return jsonify({'error': '見つかりません'}), 404
    # assigned_tables / children_info はすでにリスト型で返ってくる
    all_res = get_all_reservations(include_cancelled=False)
    ranking = aggregate_customer_ranking(all_res)
    key = customer_key(res)
    entry = next((c for c in ranking if c['key'] == key), None)
    res['is_regular_effective'] = bool(res.get('is_regular')) or (bool(entry) and entry['visit_count'] >= REGULAR_VISIT_THRESHOLD)
    return jsonify(res)


# ============================================================
# API: クイック編集（時間・席・キャンセル）
# ============================================================
@app.route('/api/reservation/<int:rid>/quick-edit', methods=['POST'])
def api_quick_edit(rid):
    data = request.get_json() or {}
    res  = get_reservation(rid)
    if not res:
        return jsonify({'ok': False, 'error': '予約が見つかりません'}), 404

    # キャンセル
    if data.get('action') == 'cancel':
        cancel_reservation(rid)
        return jsonify({'ok': True, 'action': 'cancelled'})

    # 来店受付トグル（着席済み ⇔ 未来店）
    if data.get('action') == 'toggle-visited':
        if res['status'] == 'cancelled':
            return jsonify({'ok': False, 'error': 'キャンセル済みの予約は来店受付できません'}), 400
        new_status = 'confirmed' if res['status'] == 'visited' else 'visited'
        set_reservation_status(rid, new_status)
        return jsonify({'ok': True, 'action': 'status_updated', 'status': new_status})

    # 時間 / 席を更新
    updated = dict(res)
    if 'time_slot' in data:
        updated['time_slot'] = data['time_slot']
    if 'assigned_tables' in data:
        updated['assigned_tables'] = json.dumps(data['assigned_tables'])

    # children_info がリスト型なら JSON 文字列に戻す
    if isinstance(updated.get('children_info'), list):
        updated['children_info'] = json.dumps(updated['children_info'], ensure_ascii=False)

    update_reservation(rid, updated)
    return jsonify({'ok': True, 'action': 'updated'})


# ============================================================
# API: クイック予約追加（タイムテーブルから直接作成）
# ============================================================
@app.route('/api/reservation/quick-add', methods=['POST'])
def api_quick_add():
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'ok': False, 'error': '氏名は必須です'}), 400

    adults = max(1, int(data.get('adults', 2)))
    tables = [int(t) for t in (data.get('assigned_tables') or [])]

    res_dict = {
        'date':             data.get('date', date.today().isoformat()),
        'time_slot':        data.get('time_slot', '18:00'),
        'name':             name,
        'phone':            (data.get('phone') or '').strip() or '-',
        'adults':           adults,
        'children_info':    '[]',
        'total_people':     adults,
        'duration_minutes': int(data.get('duration_minutes', 105)),
        'private_room':     0,
        'is_vip':           0,
        'is_group':         1 if adults >= 8 else 0,
        'budget_per_person': None,
        'needs_type':       None,
        'gender_male':      None,
        'gender_female':    None,
        'organizer_note':   None,
        'notes':            (data.get('notes') or '').strip() or None,
        'assigned_tables':  json.dumps(tables),
        'status':           'confirmed',
    }

    rid = save_reservation(res_dict)
    return jsonify({'ok': True, 'id': rid})


# ============================================================
# API: ウォークイン登録（事前予約なしの当日フリー来店）
# ============================================================
@app.route('/api/reservation/walkin', methods=['POST'])
def api_walkin():
    """
    現場で数秒で登録できる簡易フロー。氏名入力は不要（自動採番）で、
    人数とテーブルだけ指定すればよい。来店した瞬間の登録のため、
    保存と同時に「来店受付（visited）」状態にする。
    """
    data = request.get_json() or {}
    adults = max(1, int(data.get('adults', 2)))

    now = datetime.now()
    time_slot = f"{now.hour:02d}:{(now.minute // 15) * 15:02d}"
    today_str = date.today().isoformat()

    existing = get_reservations_by_date(today_str)
    tables = [int(t) for t in (data.get('assigned_tables') or [])]
    seat_note = None
    if not tables:
        suggestion = assign_seats(
            {'total_people': adults, 'adults': adults, 'time_slot': time_slot, 'duration_minutes': 105},
            existing
        )
        tables = suggestion['tables']
        seat_note = suggestion['note']
        if not tables:
            return jsonify({'ok': False, 'error': seat_note or '空き席が見つかりません。手動でテーブルを選んでください'}), 400

    # 名前の代わりに時刻入りの識別名を自動採番（顧客カルテ集計で他のお客様と混ざらないようにするため）
    name = f"フリー（{time_slot}来店）"

    res_dict = {
        'date':             today_str,
        'time_slot':        time_slot,
        'name':             name,
        'phone':            '-',
        'adults':           adults,
        'children_info':    '[]',
        'total_people':     adults,
        'duration_minutes': 105,
        'private_room':     0,
        'is_vip':           0,
        'is_group':         1 if adults >= 8 else 0,
        'budget_per_person': None,
        'needs_type':       None,
        'gender_male':      None,
        'gender_female':    None,
        'organizer_note':   None,
        'notes':            'ウォークイン（事前予約なし・当日フリー来店）',
        'menu_note':        None,
        'sales_amount':     None,
        'assigned_tables':  json.dumps(tables),
        'status':           'confirmed',
    }
    rid = save_reservation(res_dict)
    set_reservation_status(rid, 'visited')  # その場で来店済みとして扱う（visited_atも記録される）

    return jsonify({'ok': True, 'id': rid, 'name': name, 'time_slot': time_slot, 'tables': tables})


# ============================================================
# API: 顧客カルテ（過去履歴）呼び出し
# ============================================================
@app.route('/api/customer-history')
def api_customer_history():
    """氏名・電話番号から過去の来店履歴を返す。予約フォームからAJAXで呼ばれる。"""
    name  = request.args.get('name', '').strip()
    phone = request.args.get('phone', '').strip()
    if not name and not phone:
        return jsonify({'found': False})

    exclude_id = request.args.get('exclude_id')
    all_res = get_all_reservations(include_cancelled=False)
    history = get_customer_history(
        name, phone, all_res,
        exclude_id=int(exclude_id) if exclude_id else None,
    )
    return jsonify(history)


# ============================================================
# ページ: 分析・集計（来店頻度ランキング・売上ランキング）
# ============================================================
@app.route('/analytics')
def analytics():
    period = request.args.get('period', 'all')  # all / 30 / 90 / 365
    since_date = None
    if period != 'all':
        since_date = (date.today() - timedelta(days=int(period))).isoformat()

    all_res   = get_all_reservations(include_cancelled=False)
    customers = aggregate_customer_ranking(all_res, since_date=since_date)

    by_visits = sorted(customers, key=lambda c: (-c['visit_count'], c['days_since_last_visit']))[:50]
    by_sales  = sorted(customers, key=lambda c: -c['total_sales'])[:50]
    by_recent = sorted(customers, key=lambda c: c['days_since_last_visit'])[:50]

    total_customers  = len(customers)
    total_sales_all  = sum(c['total_sales'] for c in customers)
    total_visits_all = sum(c['visit_count'] for c in customers)
    repeat_customers = sum(1 for c in customers if c['visit_count'] >= 2)

    unmatched_sales = posimp.group_by_receipt(get_pos_sales(match_status='unmatched'))
    standalone_sales_total = sum(
        g['amount'] for g in posimp.group_by_receipt(get_pos_sales(match_status='standalone'))
    )

    return render_template(
        'analytics.html',
        period            = period,
        by_visits         = by_visits,
        by_sales          = by_sales,
        by_recent         = by_recent,
        total_customers   = total_customers,
        total_sales_all   = total_sales_all,
        total_visits_all  = total_visits_all,
        repeat_customers  = repeat_customers,
        unmatched_count   = len(unmatched_sales),
        standalone_sales_total = standalone_sales_total,
    )


# ============================================================
# マジレジ（POS）売上インポート
# ============================================================
@app.route('/pos-import')
def pos_import_page():
    return render_template('pos_import.html')


@app.route('/pos-import/preview', methods=['POST'])
def pos_import_preview():
    f = request.files.get('file')
    if not f or not f.filename:
        flash('⚠️ ファイルを選択してください', 'danger')
        return redirect(url_for('pos_import_page'))

    raw_bytes = f.read()
    if len(raw_bytes) > 10 * 1024 * 1024:  # 10MB
        flash('⚠️ ファイルサイズが大きすぎます（10MB以下にしてください）', 'danger')
        return redirect(url_for('pos_import_page'))

    headers, rows = posimp.read_table(f.filename, raw_bytes)
    if not headers or not rows:
        flash('⚠️ ファイルを読み込めませんでした（形式・文字コードをご確認ください）', 'danger')
        return redirect(url_for('pos_import_page'))

    guess = posimp.guess_columns(headers)

    return render_template(
        'pos_import_mapping.html',
        headers      = headers,
        guess        = guess,
        preview_rows = rows[:10],
        row_count    = len(rows),
        filename     = f.filename,
        file_b64     = base64.b64encode(raw_bytes).decode('ascii'),
        target_fields = [
            ('sale_date',      '会計日',   True),
            ('sale_time',      '会計時刻', False),
            ('receipt_no',     '伝票番号', False),
            ('table_no',       '卓番号',   False),
            ('party_size',     '客数',     False),
            ('amount',         '金額',     True),
            ('item_name',      '商品名',   False),
            ('customer_name',  'お客様名', False),
            ('phone',          '電話番号', False),
            ('payment_method', '支払方法', False),
        ],
    )


@app.route('/pos-import/commit', methods=['POST'])
def pos_import_commit():
    file_b64 = request.form.get('file_b64', '')
    filename = request.form.get('filename', 'upload.csv')
    try:
        raw_bytes = base64.b64decode(file_b64)
    except Exception:
        flash('⚠️ アップロードデータの読み込みに失敗しました。もう一度アップロードしてください', 'danger')
        return redirect(url_for('pos_import_page'))

    headers, rows = posimp.read_table(filename, raw_bytes)
    mapping = {field: (request.form.get(f'map_{field}') or None) for field in posimp.TARGET_FIELDS}

    missing = [f for f in posimp.REQUIRED_FIELDS if not mapping.get(f)]
    if missing:
        flash(f"⚠️ 必須項目が未設定です: {'・'.join(missing)}", 'danger')
        guess = posimp.guess_columns(headers)
        return render_template(
            'pos_import_mapping.html',
            headers=headers, guess=guess, preview_rows=rows[:10], row_count=len(rows),
            filename=filename, file_b64=file_b64,
            target_fields=[
                ('sale_date', '会計日', True), ('sale_time', '会計時刻', False),
                ('receipt_no', '伝票番号', False), ('table_no', '卓番号', False),
                ('party_size', '客数', False), ('amount', '金額', True),
                ('item_name', '商品名', False), ('customer_name', 'お客様名', False),
                ('phone', '電話番号', False), ('payment_method', '支払方法', False),
            ],
        )

    import_batch = f"{filename}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    normalized_rows, error_count = posimp.normalize_rows(rows, mapping, import_batch)
    inserted_count = insert_pos_sales(normalized_rows)

    pos_groups = posimp.group_by_receipt(normalized_rows)
    dates = sorted(set(g['sale_date'] for g in pos_groups))
    reservations_by_date = {d: get_reservations_by_date(d) for d in dates}
    matched_groups = match_pos_sale_groups(pos_groups, reservations_by_date)

    matched = [g for g in matched_groups if g['matched_reservation_id']]
    unmatched = [g for g in matched_groups if not g['matched_reservation_id']]
    for g in matched:
        apply_pos_sale_to_reservation(g['matched_reservation_id'], g['amount'], g['menu_items'])
        set_pos_sale_group_match(g['receipt_key'], g['matched_reservation_id'], 'matched')

    return render_template(
        'pos_import_result.html',
        row_count      = len(rows),
        inserted_count = inserted_count,
        error_count    = error_count,
        group_count    = len(pos_groups),
        matched        = matched,
        unmatched      = unmatched,
    )


@app.route('/pos-import/unmatched')
def pos_import_unmatched():
    groups = posimp.group_by_receipt(get_pos_sales(match_status='unmatched'))
    groups.sort(key=lambda g: (g['sale_date'], g['sale_time'] or ''), reverse=True)
    return render_template('pos_unmatched.html', groups=groups)


@app.route('/pos-sale/link', methods=['POST'])
def pos_sale_link():
    receipt_key = request.form.get('receipt_key', '')
    action      = request.form.get('action', '')

    rows = get_pos_sales()
    group_rows = [r for r in rows if r['receipt_key'] == receipt_key]
    if not group_rows:
        flash('⚠️ 対象のPOS売上が見つかりません', 'danger')
        return redirect(url_for('pos_import_unmatched'))

    amount = sum(r['amount'] for r in group_rows)
    items  = [r['item_name'] for r in group_rows if r['item_name']]

    if action == 'standalone':
        set_pos_sale_group_match(receipt_key, None, 'standalone')
        flash('✅ 単体売上（予約なし）として登録しました', 'success')
    else:
        rid = request.form.get('reservation_id')
        if not rid:
            flash('⚠️ 紐付ける予約を選択してください', 'danger')
            return redirect(url_for('pos_import_unmatched'))
        apply_pos_sale_to_reservation(int(rid), amount, items)
        set_pos_sale_group_match(receipt_key, int(rid), 'manual')
        flash('✅ 予約に手動で紐付けました', 'success')

    return redirect(url_for('pos_import_unmatched'))


@app.route('/api/reservations-by-date')
def api_reservations_by_date():
    target_date = request.args.get('date', '')
    if not target_date:
        return jsonify([])
    rows = get_reservations_by_date(target_date)
    return jsonify([
        {'id': r['id'], 'name': r['name'], 'time_slot': r['time_slot'], 'total_people': r['total_people']}
        for r in rows
    ])


# ============================================================
# バックアップ: CSV全件エクスポート
# ============================================================
@app.route('/backup/csv')
def backup_csv():
    """全予約データをCSVでダウンロード"""
    with engine.connect() as con:
        rows = con.execute(
            text("SELECT * FROM reservations ORDER BY date, time_slot")
        ).mappings().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'ID', '日付', '時間', '氏名', '電話番号',
        '大人', '子供', '合計人数', '利用時間(分)',
        '個室', 'VIP', '団体', '予算/人', 'ニーズ',
        '男性', '女性', '幹事メモ', '備考', 'メニュー', '売上',
        'テーブル', 'ステータス', '来店受付日時', '登録日時', '更新日時',
    ])
    for r in rows:
        d = dict(r)
        writer.writerow([
            d['id'], d['date'], d['time_slot'], d['name'], d['phone'],
            d['adults'],
            len(json.loads(d['children_info'] or '[]')),
            d['total_people'], d['duration_minutes'],
            '○' if d['private_room'] else '',
            '○' if d['is_vip'] else '',
            '○' if d['is_group'] else '',
            d['budget_per_person'] or '',
            d['needs_type'] or '',
            d['gender_male'] or '', d['gender_female'] or '',
            d['organizer_note'] or '', d['notes'] or '',
            d.get('menu_note') or '', d.get('sales_amount') or '',
            d['assigned_tables'], d['status'], d.get('visited_at') or '',
            d['created_at'], d['updated_at'],
        ])

    fname = f"kiriniya_backup_{date.today().isoformat()}.csv"
    return Response(
        '﻿' + output.getvalue(),
        mimetype='text/csv; charset=utf-8-sig',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


# ============================================================
# バックアップ: 顧客集計 CSV エクスポート
# ============================================================
@app.route('/backup/customers/csv')
def backup_customers_csv():
    """顧客ごとの来店回数・累計売上をCSVでダウンロード"""
    all_res   = get_all_reservations(include_cancelled=False)
    customers = aggregate_customer_ranking(all_res)
    customers.sort(key=lambda c: -c['visit_count'])

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['氏名', '電話番号', '来店回数', '累計売上', '平均単価/回',
                      '初回来店', '前回来店', '前回からの経過日数', '最終来店受付日時'])
    for c in customers:
        writer.writerow([
            c['name'], c['phone'], c['visit_count'], c['total_sales'], c['avg_sales'],
            c['first_visit_date'], c['last_visit_date'], c['days_since_last_visit'],
            c.get('last_visited_at') or '',
        ])

    fname = f"kiriniya_customers_{date.today().isoformat()}.csv"
    return Response(
        '﻿' + output.getvalue(),
        mimetype='text/csv; charset=utf-8-sig',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


# ============================================================
# バックアップ: POS売上明細 CSV エクスポート
# ============================================================
@app.route('/backup/pos/csv')
def backup_pos_csv():
    """マジレジ等からインポートしたPOS売上明細を全件CSVでダウンロード"""
    rows = get_pos_sales()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['日付', '時刻', '伝票番号', '卓番号', '客数', '商品名', '金額',
                      'お客様名', '電話番号', '突合ステータス', '紐付け予約ID', 'インポート回次'])
    for r in rows:
        writer.writerow([
            r['sale_date'], r['sale_time'] or '', r['receipt_no'] or '', r['table_no'] or '',
            r['party_size'] or '', r['item_name'] or '', r['amount'],
            r['customer_name'] or '', r['phone'] or '',
            r['match_status'], r['matched_reservation_id'] or '', r['import_batch'],
        ])

    fname = f"kiriniya_pos_sales_{date.today().isoformat()}.csv"
    return Response(
        '﻿' + output.getvalue(),
        mimetype='text/csv; charset=utf-8-sig',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


# ============================================================
# バックアップ: Excel(XLSX) 全件エクスポート（予約全件 + 顧客集計 + POS売上明細）
# ============================================================
@app.route('/backup/xlsx')
def backup_xlsx():
    all_res   = get_all_reservations(include_cancelled=True)
    customers = aggregate_customer_ranking([r for r in all_res if r['status'] != 'cancelled'])
    customers.sort(key=lambda c: -c['visit_count'])
    pos_rows  = get_pos_sales()

    wb = Workbook()

    ws1 = wb.active
    ws1.title = '予約全件'
    ws1.append([
        'ID', '日付', '時間', '氏名', '電話番号',
        '大人', '子供', '合計人数', '利用時間(分)',
        '個室', 'VIP', '団体', '予算/人', 'ニーズ',
        '男性', '女性', '幹事メモ', '備考', 'メニュー', '売上',
        'テーブル', 'ステータス', '来店受付日時', '登録日時', '更新日時',
    ])
    for r in all_res:
        ws1.append([
            r['id'], r['date'], r['time_slot'], r['name'], r['phone'],
            r['adults'], len(r['children_info']), r['total_people'], r['duration_minutes'],
            '○' if r['private_room'] else '', '○' if r['is_vip'] else '', '○' if r['is_group'] else '',
            r.get('budget_per_person') or '', r.get('needs_type') or '',
            r.get('gender_male') or '', r.get('gender_female') or '',
            r.get('organizer_note') or '', r.get('notes') or '',
            r.get('menu_note') or '', r.get('sales_amount') or '',
            '・'.join(str(t) for t in r['assigned_tables']), r['status'], r.get('visited_at') or '',
            r['created_at'], r['updated_at'],
        ])

    ws2 = wb.create_sheet('顧客集計')
    ws2.append(['氏名', '電話番号', '来店回数', '累計売上', '平均単価/回',
                '初回来店', '前回来店', '前回からの経過日数', '最終来店受付日時'])
    for c in customers:
        ws2.append([
            c['name'], c['phone'], c['visit_count'], c['total_sales'], c['avg_sales'],
            c['first_visit_date'], c['last_visit_date'], c['days_since_last_visit'],
            c.get('last_visited_at') or '',
        ])

    ws3 = wb.create_sheet('POS売上明細')
    ws3.append(['日付', '時刻', '伝票番号', '卓番号', '客数', '商品名', '金額',
                'お客様名', '電話番号', '突合ステータス', '紐付け予約ID', 'インポート回次'])
    for r in pos_rows:
        ws3.append([
            r['sale_date'], r['sale_time'] or '', r['receipt_no'] or '', r['table_no'] or '',
            r['party_size'] or '', r['item_name'] or '', r['amount'],
            r['customer_name'] or '', r['phone'] or '',
            r['match_status'], r['matched_reservation_id'] or '', r['import_batch'],
        ])

    for ws in (ws1, ws2, ws3):
        for col_cells in ws.columns:
            length = max((len(str(cell.value)) for cell in col_cells if cell.value is not None), default=8)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"kiriniya_full_{date.today().isoformat()}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


# ============================================================
# API: 同期ポーリング用（最終更新タイムスタンプを返す）
# ============================================================
@app.route('/api/last-change')
def api_last_change():
    """
    予約テーブルの最終更新日時をDBから直接取得。
    最後に更新された予約がキャンセルだった場合は、急なキャンセルをその場で
    全端末に知らせられるよう、氏名・時間などの情報も併せて返す。
    """
    try:
        with engine.connect() as con:
            row = con.execute(text(
                "SELECT date, time_slot, name, status, updated_at FROM reservations "
                "ORDER BY updated_at DESC LIMIT 1"
            )).mappings().first()
        if not row:
            return jsonify({'ts': ''})
        return jsonify({
            'ts':             row['updated_at'] or '',
            'last_status':    row['status'],
            'last_name':      row['name'],
            'last_time_slot': row['time_slot'],
            'last_date':      row['date'],
        })
    except Exception:
        return jsonify({'ts': ''})


# ============================================================
# ページ: 接続情報（タブレット共有用 QR・URL）
# ============================================================
@app.route('/share')
def share_page():
    return render_template('share.html', local_ip=LOCAL_IP)


@app.route('/init-db')
def init_db_route():
    init_db()
    return '✅ DB初期化完了'


# ============================================================
# フォームパーサー
# ============================================================
def _parse_form(req, edit_id=None) -> tuple[dict, str | None]:
    """request.form → dict 変換。エラー時は (None, エラーメッセージ)"""
    f = req.form
    try:
        child_ages    = req.form.getlist('child_age[]')
        children_info = [{"age": int(a)} for a in child_ages if a.strip().isdigit()]
        adults        = int(f.get('adults', 1))
        total         = adults + len(children_info)
        special_tags  = [t for t in req.form.getlist('special_tags[]') if t in SPECIAL_TAG_CHOICES]

        d = {
            'date':              f['date'],
            'time_slot':         f['time_slot'],
            'name':              f['name'].strip(),
            'phone':             f['phone'].strip(),
            'adults':            adults,
            'children_info':     json.dumps(children_info, ensure_ascii=False),
            'total_people':      total,
            'duration_minutes':  int(f.get('duration_minutes', 105)),
            'private_room':      1 if 'private_room' in f else 0,
            'is_vip':            1 if 'is_vip' in f else 0,
            'is_regular':        1 if 'is_regular' in f else 0,
            'is_group':          1 if total >= 8 else 0,
            'budget_per_person': int(f['budget_per_person']) if f.get('budget_per_person', '').strip() else None,
            'needs_type':        f.get('needs_type') or None,
            'gender_male':       int(f['gender_male'])   if f.get('gender_male',   '').strip().isdigit() else None,
            'gender_female':     int(f['gender_female'])  if f.get('gender_female', '').strip().isdigit() else None,
            'organizer_note':    f.get('organizer_note', '').strip() or None,
            'notes':             f.get('notes', '').strip() or None,
            'menu_note':         f.get('menu_note', '').strip() or None,
            'sales_amount':      int(f['sales_amount']) if f.get('sales_amount', '').strip() else None,
            'special_tags':      json.dumps(special_tags, ensure_ascii=False),
            'assigned_tables':   '[]',
            'status':            'confirmed',
        }
        if edit_id:
            d['id'] = edit_id

        if not d['name'] or not d['phone']:
            return None, '氏名・電話番号は必須です'
        return d, None
    except (KeyError, ValueError) as e:
        return None, str(e)

def _parse_confirmed_tables(raw: str) -> list[int]:
    """フォームの confirmed_tables フィールドをパース。"""
    try:
        data = json.loads(raw or '[]')
        return [int(t) for t in data if str(t).isdigit() or isinstance(t, int)]
    except Exception:
        return []

# ============================================================
# エントリポイント
# ============================================================
if __name__ == '__main__':
    import sys
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    if hasattr(sys.stderr, 'reconfigure'):
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    init_db()
    print()
    print("=" * 50)
    print("  Kirin-ya Yoyaku System  Start")
    print("=" * 50)
    print("  http://localhost:5000")
    print("  Ctrl+C de teishi")
    print()
    app.run(debug=True, port=5000, host='0.0.0.0', threaded=True)
