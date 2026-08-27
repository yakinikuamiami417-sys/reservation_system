# pos_import.py — マジレジ（POSレジ）売上CSV/Excelのインポート処理
# キリン屋 予約管理システム
#
# POSの出力形式は店舗設定により列名・行の粒度（伝票単位／商品明細単位）が異なるため、
# 列名を決め打ちせず「候補ヘッダーからの自動検出 → 人間による確認・修正」を前提にする。
import csv, io, re
from datetime import datetime

from openpyxl import load_workbook

# ============================================================
# 列の自動検出候補
# ============================================================
# ターゲットフィールド名 → CSV側にありそうなヘッダー名の候補（先頭からの優先順）
COLUMN_CANDIDATES = {
    'sale_date':      ['会計日', '来店日', '営業日', '伝票日付', '取引日', '日付'],
    'sale_time':      ['会計時刻', '会計時間', '伝票時刻', '取引時刻', '時刻', '時間'],
    'receipt_no':     ['伝票番号', '伝票No', '伝票NO', 'レシート番号', '取引番号', '会計番号'],
    'table_no':       ['卓番号', '卓番', 'テーブル番号', 'テーブル', '座席番号', '席番号'],
    'party_size':     ['客数', '人数', '来店人数', '組人数'],
    'amount':         ['合計金額', 'お会計金額', '売上金額', '税込合計', '合計', '金額', 'お会計'],
    'item_name':      ['商品名', '品名', 'メニュー', '商品', 'メニュー名'],
    'customer_name':  ['お客様名', 'お名前', '氏名', '顧客名', '会員名'],
    'phone':          ['電話番号', '電話', 'TEL', 'ＴＥＬ'],
    'payment_method':  ['支払方法', '決済方法', '支払種別', '支払'],
}

REQUIRED_FIELDS = ['sale_date', 'amount']
TARGET_FIELDS = list(COLUMN_CANDIDATES.keys())


def guess_columns(headers: list) -> dict:
    """CSVヘッダー一覧から、ターゲットフィールドごとに最も近そうな列名を推測する。"""
    normalized = {h: h.strip() for h in headers}
    guess = {}
    for field, candidates in COLUMN_CANDIDATES.items():
        found = None
        for cand in candidates:
            for h, hn in normalized.items():
                if hn == cand:
                    found = h
                    break
            if found:
                break
        if not found:
            # 部分一致でも探す（例: '合計金額(税込)' のような列名対策）
            for cand in candidates:
                for h, hn in normalized.items():
                    if cand in hn:
                        found = h
                        break
                if found:
                    break
        guess[field] = found
    return guess


# ============================================================
# ファイル読み込み（CSV / Excel、文字コード自動判定）
# ============================================================
def read_table(filename: str, raw_bytes: bytes):
    """
    アップロードされたファイルを読み込み、(headers, rows) を返す。
    rows は「ヘッダー名 → 値（文字列）」の dict のリスト。
    CSVは日本語POSに多い Shift_JIS(CP932) を優先的に試し、ダメなら UTF-8 系にフォールバックする。
    """
    ext = (filename or '').lower().rsplit('.', 1)[-1] if '.' in (filename or '') else ''

    if ext in ('xlsx', 'xlsm'):
        wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], []
        headers = [str(h).strip() if h is not None else f'列{i+1}' for i, h in enumerate(header_row)]
        rows = []
        for r in rows_iter:
            if r is None or all(v is None for v in r):
                continue
            rows.append({headers[i]: ('' if v is None else str(v)) for i, v in enumerate(r) if i < len(headers)})
        return headers, rows

    # CSV: 日本語POSはShift_JIS(CP932)出力が多いため、まずcp932を試す
    text_data = None
    for enc in ('cp932', 'utf-8-sig', 'utf-8'):
        try:
            text_data = raw_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text_data is None:
        text_data = raw_bytes.decode('utf-8', errors='replace')

    reader = csv.reader(io.StringIO(text_data))
    rows_raw = [row for row in reader if any(cell.strip() for cell in row)]
    if not rows_raw:
        return [], []
    headers = [h.strip() for h in rows_raw[0]]
    rows = []
    for r in rows_raw[1:]:
        row_dict = {}
        for i, h in enumerate(headers):
            row_dict[h] = r[i].strip() if i < len(r) else ''
        rows.append(row_dict)
    return headers, rows


# ============================================================
# 値のパース
# ============================================================
_DATE_FORMATS = ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y年%m月%d日']

def parse_date(raw: str):
    if not raw:
        return None
    raw = raw.strip()
    # Excelのシリアル値や日時が混ざっている場合、先頭の日付部分だけ拾う
    m = re.match(r'(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})', raw)
    if m:
        y, mo, d = m.groups()
        try:
            return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
        except ValueError:
            return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def parse_time(raw: str):
    if not raw:
        return None
    raw = raw.strip()
    m = re.match(r'(\d{1,2})[:：時](\d{1,2})', raw)
    if m:
        h, mi = m.groups()
        try:
            return f"{int(h):02d}:{int(mi):02d}"
        except ValueError:
            return None
    return None


def parse_amount(raw: str):
    if raw is None:
        return 0
    cleaned = re.sub(r'[^\d\-]', '', str(raw))
    try:
        return int(cleaned) if cleaned not in ('', '-') else 0
    except ValueError:
        return 0


def parse_int(raw: str):
    if raw is None or str(raw).strip() == '':
        return None
    cleaned = re.sub(r'[^\d]', '', str(raw))
    return int(cleaned) if cleaned else None


# ============================================================
# 日計の合計行・小計行の検出
# ============================================================
# POSの日計CSV/Excelはデータ行の末尾に「合計」「総合計」等のサマリー行が
# 付与されることが多い。これを個別の会計（顧客データ）として取り込んでしまうと
# 全売上合算額が1件の巨大な伝票として登録されてしまうため、事前に除外する。
# ラベルがどの列（卓番号欄・伝票番号欄・商品名欄など）に入っていても検出できるよう、
# マッピング先の列に関わらず行内の全セル値をチェックする。
_SUMMARY_ROW_KEYWORDS = {
    '合計', '総合計', '小計', '日計', '総計', '合計金額', '総合計金額',
    'TOTAL', 'Total', 'total', '合計欄',
}

def is_summary_row(raw_row: dict) -> bool:
    """行内のいずれかのセルが集計ラベルと完全一致する場合、合計・小計行とみなす。"""
    for v in raw_row.values():
        if v is None:
            continue
        if str(v).strip() in _SUMMARY_ROW_KEYWORDS:
            return True
    return False


# ============================================================
# 行の正規化 & 伝票単位へのグルーピング
# ============================================================
def normalize_rows(rows: list, mapping: dict, import_batch: str):
    """
    生の行データを、DB登録用の正規化行（1行=1明細）に変換する。
    mapping: {target_field: csv_header_name}
    戻り値: (normalized_rows, error_count, skipped_summary_count)
    """
    normalized = []
    errors = 0
    skipped_summary = 0
    for idx, raw in enumerate(rows):
        def get(field):
            col = mapping.get(field)
            return raw.get(col, '') if col else ''

        # 日計の合計行・小計行は個別の会計ではないため、顧客データとして取り込まない
        if is_summary_row(raw):
            skipped_summary += 1
            continue

        sale_date = parse_date(get('sale_date'))
        amount    = parse_amount(get('amount'))
        if not sale_date:
            errors += 1
            continue

        sale_time  = parse_time(get('sale_time'))
        receipt_no = (get('receipt_no') or '').strip() or None
        table_no   = (get('table_no') or '').strip() or None
        # グルーピングキーの優先順位:
        #   ① 伝票番号（最も確実。同じ卓の複数回転でも伝票番号が異なれば別会計として扱われる）
        #   ② 日付+卓番号+会計時刻（伝票番号が無い場合。時刻が取れていれば回転ごとに別キーになる）
        #   ③ 日付+卓番号+行番号（伝票番号も会計時刻も取れない場合のフォールバック）
        #      ※ 時刻情報が無いまま「日付+卓番号」だけでキーを作ると、同じ卓の複数回転が
        #        すべて同一キーに収束し、金額が誤って合算されてしまうため、行番号で必ず
        #        ユニークにする（＝伝票番号も時刻も無い行は決して自動合算しない）。
        if receipt_no:
            receipt_key = receipt_no
        elif sale_time:
            receipt_key = f"{sale_date}|{table_no or '?'}|{sale_time}"
        else:
            receipt_key = f"{sale_date}|{table_no or '?'}|row{idx}"

        normalized.append({
            'import_batch':   import_batch,
            'sale_date':      sale_date,
            'sale_time':      sale_time,
            'receipt_no':     receipt_no,
            'table_no':       table_no,
            'party_size':     parse_int(get('party_size')),
            'amount':         amount,
            'item_name':      (get('item_name') or '').strip() or None,
            'customer_name':  (get('customer_name') or '').strip() or None,
            'phone':          (get('phone') or '').strip() or None,
            'payment_method': (get('payment_method') or '').strip() or None,
            'receipt_key':    receipt_key,
            'raw_row':        raw,
        })
    return normalized, errors, skipped_summary


def group_by_receipt(normalized_rows: list) -> list:
    """
    明細行を伝票（receipt_key）単位にまとめる。
    商品明細が複数行に分かれている場合はamountを合算、item_nameを連結する。
    """
    groups = {}
    for r in normalized_rows:
        key = r['receipt_key']
        g = groups.setdefault(key, {
            'receipt_key':   key,
            'sale_date':     r['sale_date'],
            'sale_time':     r['sale_time'],
            'receipt_no':    r['receipt_no'],
            'table_no':      r['table_no'],
            'party_size':    r['party_size'],
            'amount':        0,
            'menu_items':    [],
            'customer_name': r['customer_name'],
            'phone':         r['phone'],
            'row_count':     0,
        })
        g['amount'] += r['amount']
        g['row_count'] += 1
        if r['item_name']:
            g['menu_items'].append(r['item_name'])
        # 明細行によって時刻・卓番号が空のものがあれば埋める
        g['sale_time']  = g['sale_time']  or r['sale_time']
        g['table_no']   = g['table_no']   or r['table_no']
        g['party_size'] = g['party_size'] or r['party_size']
        g['customer_name'] = g['customer_name'] or r['customer_name']
        g['phone']          = g['phone']          or r['phone']
    return list(groups.values())
